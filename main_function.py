import itertools
import pandas as pd
import os
import numpy as np
import csv
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.worksheet.pagebreak import Break, PageBreak
def break_list(f,min_deffrence,total_stu):
    x1=[]
    x2=[]
    for i in f:
        for j in f:
            if abs(sum(i)-sum(j))==min_deffrence and (sum(i)+sum(j))==total_stu and i!=j and len((set(i).intersection(set(j))))==0:
                #print(i,sum(i),j,sum(j),sum(total_stu))
                x1=i
                x2=j
                break
        if len(x1)>0:
            return x1,x2
            break
def findMinAbsDiff(S, n, S1=0, S2=0):
	# Base case: if the list becomes empty, return the absolute
	# difference between both sets
	if n < 0:
		return abs(S1 - S2)
	# Case 1. Include the current item in subset `S1` and recur
	# for the remaining items `n-1`
	inc = findMinAbsDiff(S, n - 1, S1 + S[n], S2)
	# Case 2. Exclude the current item from subset `S1` and recur for
	# the remaining items `n-1`
	exc = findMinAbsDiff(S, n - 1, S1, S2 + S[n])
	return min(inc, exc)

def findsubsets(s):
    l=[]
    for n in range(1,len(s)+1):
        l.extend(list(itertools.combinations(s, n)))
    return l
def get_class_list(x,y,classes):
    list1=[]
    list2=[]
    key=classes.keys()
    for itm in x:
        for k in key:
            if classes[k]==itm:
                list1.append(k)
                classes[k]=0
    for itm in y:
        for k in key:
            if classes[k]==itm:
                list2.append(k)
                classes[k]=0
    return list1,list2

def assign_group_to_class(df,class_list,group=1):
        temp=pd.DataFrame()
        for c in class_list:
                temp=pd.concat([temp,df[df['Class']==c]])
                temp['Group']=group
        return temp
def genrate_id(df):
    t=[]
    for i in df.iterrows():
                     k=df[df['PR']==i[1][0]]
                     t.append(str(list(k['PR'])[0])+"-"+str(list(k['Class'])[0])+"-"+(str(list(k['Section'])[0])[-1:-4:-1])[-1:-4:-1]+'-'+str(list(k['Gender'])[0]))
    df['ID']=t
    return df
def genrate_roll_no(df):
    section=df.Section.unique()
    df2=pd.DataFrame()
    for s in section:
        temp=pd.DataFrame(df[df['Section']==s])
        temp.sort_values(by=['Name'],inplace=True)
        temp['RN']=[i for i in range(1,len(temp)+1)]
        df2=pd.concat([df2,temp])
    return df2
def arrange_alternate(group1,group2,x):
    dfix=len(group1)
    dfx=len(group2)
    df_final=pd.DataFrame()
    #simple_df=pd.DataFrame()
    #alternate_df=pd.DataFrame()
    dff=pd.DataFrame()
    for i in range(0,max(dfix,dfx),x//2):
        first_df=pd.DataFrame(group1,
                         columns=['PR','Name','RN','Class','Section',"ID"])[i:i+x//2]
    #print(dfix_14)
        second_df=pd.DataFrame(group2,
                        columns=['PR','Name','RN','Class','Section',"ID"])[i:i+x//2]
        df_final=pd.concat([df_final,first_df,second_df])
        j,k=0,0
        for i in range(x):
        #arrange alternate
            if i%2==0:
                dff=pd.concat([dff,first_df[j:j+1]])
                j+=1
            else:
                dff=pd.concat([dff,second_df[k:k+1]])
                k+=1
    return dff
    '''for i in range(0,len(df_final),x):
        df=pd.DataFrame(df_final[i:i+x])
        df["Room"]=(i//x)+1
        simple_df=pd.concat([simple_df,df])'''
    
def room_alotment(df,n_seat,room):
    alternate_df=pd.DataFrame()
    for i in range(0,len(df),n_seat):
        df3=pd.DataFrame(df[i:i+n_seat])
        df3["Room"]=(i//n_seat)+1#room[i]
        alternate_df=pd.concat([alternate_df,df3])
    return alternate_df
def genrate_attendance_sheet(att_df,date,file_name=''):
    wb = openpyxl.Workbook() 
    sheet = wb.active
    start_row=1
    room=list(att_df['Room'].unique())
    date=[]
    for r in room:
        d=pd.DataFrame(att_df[att_df['Room']==r],columns=['PR','Name','RN','Class','Section','ID','Room'])
        d['Name']=d['Name'].str.upper()
        d=d.sort_values(by=['Class', 'Section','Name'])
        d.reset_index(inplace = True)
        c_name='ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        col=columns=['PR','Name','RN','Class','Section','ID','25-02-22','26-02-22','28-02-22','01-03-22','02-03-22','03-03-22','04-03-22']
        sheet.merge_cells('A'+str(start_row)+':'+c_name[len(col)-1]+str(start_row))
        sheet.cell(row = start_row, column = 1).value='Room No-'+str(r)+" (PreBoard-Feb-Mar-22)"
        sheet.cell(row = start_row, column = 1).font=Font(size = 18,bold=True)
        sheet.cell(row = start_row, column = 1).alignment=Alignment(horizontal='center')
        start_row+=1
    
        k=1
        for c in col:
            sheet.cell(row=start_row,column=k).value=c
            sheet.cell(row = start_row, column = k).font=Font(size = 12,bold=True)
            sheet.cell(row = start_row, column = k).alignment=Alignment(horizontal='center')
            k+=1
        for i in range(0,len(d)):
            start_row+=1
            #print(len(sheet['A']),'start_row',start_row,"room_no",r)
            sheet.cell(row=start_row,column=1).value=d.loc[i][1]
            sheet.cell(row=start_row,column=2).value=d.loc[i][2]
            sheet.cell(row=start_row,column=3).value=d.loc[i][3]
            sheet.cell(row=start_row,column=4).value=d.loc[i][4]
            sheet.cell(row=start_row,column=5).value=d.loc[i][5]
            sheet.cell(row=start_row,column=6).value=d.loc[i][6]
            #sheet.cell(row=start_row,column=7).value=d.loc[i][7]
        start_row+=1
        sheet.merge_cells('A'+str(start_row)+':F'+str(start_row))
        sheet.cell(row=start_row,column=1).value='Total Present'
        sheet.cell(row = start_row, column = 1).font=Font(size = 10,bold=True)
        sheet.cell(row = start_row, column = 1).alignment=Alignment(horizontal='right')
        start_row+=1
        sheet.merge_cells('A'+str(start_row)+':F'+str(start_row))
        sheet.cell(row=start_row,column=1).value='Total Absent'
        sheet.cell(row = start_row, column = 1).font=Font(size = 10,bold=True)
        sheet.cell(row = start_row, column = 1).alignment=Alignment(horizontal='right')
        start_row+=1
        sheet.merge_cells('A'+str(start_row)+':F'+str(start_row))
        sheet.cell(row=start_row,column=1).value='Signature'
        sheet.cell(row = start_row, column = 1).font=Font(size = 10,bold=True)
        sheet.cell(row = start_row, column = 1).alignment=Alignment(horizontal='right')
        start_row=len(sheet['A'])+1
    
    wb.save(file_name+' Attendance sheet.xlsx')
def genrate_seat_plan(df,n_seat,n_row,file_name='Alternate '):
    wb = openpyxl.Workbook() 
    sheet = wb.active 
    #df= alternate_dataframe
    n_col=n_seat//n_row
    room=list(df['Room'].unique())
    start_row=1
    for r in room:
        l=[np.NAN]*(n_seat)
        l2=[np.NAN]*(n_seat//2)
        xx=list(pd.DataFrame(df[df['Room']==r],columns=['ID','RN','Class'])['ID'])
        #num=len(xx)
        d=pd.DataFrame(df[df['Room']==r],columns=['ID','RN','Class','Section'])
        '''row1=list(pd.DataFrame(df[0][df[0]['Room']==r],columns=['ID','RN','Class'])['ID'])[0:num//2]
        #rn1=list(pd.DataFrame(df[0][df[0]['Room']==r],columns=['ID','RN','Class'])['RN'])[0:num//2]
        row2=list(pd.DataFrame(df[0][df[0]['Room']==r],columns=['ID','RN','Class'])['ID'])[num//2:]
        #rn2=list(pd.DataFrame(df[0][df[0]['Room']==r],columns=['ID','RN','Class'])['RN'])[0:num//2:]
        #print(len(row1)+len(row2))'''
        for i in xx:
            if xx.index(i)<len(l):
                l[xx.index(i)]=str(i)+' Roll No. '+str(int(list(df[df['ID']==i]['RN'])[0]))#"R.No"+str(r)
                #l1[row1.index(i)]=i+' R.No:'+str(rn1[row1.index(i)])
        '''for j in row2:
            if row2.index(j)<len(l2):
                #print(type(list(pd.DataFrame(df[2][df[2]['ID']==j],columns=['ID','RN'])['RN'])[0]))
                l2[row2.index(j)]=str(j)+'Roll No. '+str(int(list(df[0][df[0]['ID']==j]['RN'])[0]))
        #print(l1,l2)'''
        dict1={}
        for i in range(0,n_row):
            dict1['Row'+str(i+1)]=l[i*n_col:n_col*(i+1)]
        #print(dict1)
        #ddd={'Row1':l1[0:num//n_row],'Row3':l2[0:num//n_row],'Row2':l1[num//n_row:num//2],'Row4':l2[num//n_row:num//2]}
        #print(ddd)
        #dict1={'Row1':l1[0:num//4],'Row3':l2[0:num//4],'Row2':l1[num//4:num//2],'Row4':l2[num//4:num//2]}
        df1=pd.DataFrame(dict1,columns=dict1.keys(),index=np.arange(1,(n_seat//n_row)+1))
        #print(df1)
        sheet.merge_cells('A'+str(start_row)+':'+str(chr(64+n_row))+str(start_row))
        sheet.cell(row = start_row, column = 1).value='Room No-'+str(r)+" (PreBoard-Feb-Mar-22)"
        #sheet.cell(row=start_row,column=1).width=sheet.cell(row=start_row,column=1).value
        sheet.cell(row = start_row, column = 1).font=Font(size = 24,bold=True)
        sheet.cell(row = start_row, column = 1).alignment=Alignment(horizontal='center')
        start_row+=1
        for k in range(len(dict1)):
            sheet.cell(row=start_row,column=k+1).value='Row '+str(k+1)
            sheet.cell(row = start_row, column = k+1).alignment=Alignment(horizontal='center')
            sheet.cell(row = start_row, column = k+1).font=Font(size = 16,bold=True)
        '''sheet.cell(row=start_row,column=2).value='Row 2'
        sheet.cell(row = start_row, column = 2).alignment=Alignment(horizontal='center')
        sheet.cell(row = start_row, column = 2).font=Font(size = 16,bold=True)
        sheet.cell(row=start_row,column=3).value='Row 3'
        sheet.cell(row = start_row, column = 3).alignment=Alignment(horizontal='center')
        sheet.cell(row = start_row, column = 3).font=Font(size = 16,bold=True)
        sheet.cell(row=start_row,column=4).value='Row 4'
        sheet.cell(row = start_row, column = 4).alignment=Alignment(horizontal='center')
        sheet.cell(row = start_row, column = 4).font=Font(size = 16,bold=True)  '''  
        for i in range(1,len(df1)+1):
            start_row+=1
            for col in range(n_row):
                sheet.cell(row=start_row,column=col+1).value=df1.loc[i][col]
            '''sheet.cell(row=start_row,column=2).value=df1.loc[i][1]
            sheet.cell(row=start_row,column=3).value=df1.loc[i][2]
            sheet.cell(row=start_row,column=4).value=df1.loc[i][3]'''
        total=d.pivot_table(index=['Class','Section'],values=['ID'],aggfunc='count')
        start_row+=1
        sheet.cell(row=start_row,column=1).value='Class'
        sheet.cell(row = start_row, column = 1).font=Font(size = 13,bold=True)
        sheet.cell(row=start_row,column=2).value='Section'
        sheet.cell(row = start_row, column = 2).font=Font(size = 13,bold=True)
        sheet.cell(row=start_row,column=3).value='No Of Student'
        sheet.cell(row = start_row, column = 3).font=Font(size = 13,bold=True)
        for x in total.ID.iteritems():        
            start_row+=1
            sheet.cell(row=start_row,column=1).value=x[0][0]
            sheet.cell(row = start_row, column = 1).font=Font(size = 12,bold=True)
            sheet.cell(row=start_row,column=2).value=x[0][1]
            sheet.cell(row = start_row, column = 2).font=Font(size = 12,bold=True)
            sheet.cell(row=start_row,column=3).value=x[1]
            sheet.cell(row = start_row, column = 3).font=Font(size = 12,bold=True)
    
        next_page_horizon, next_page_vertical = sheet.page_breaks
        next_page_horizon.append(Break(start_row))
        start_row=len(sheet['A'])+1
    wb.save(file_name+' Seating Plan.xlsx')
