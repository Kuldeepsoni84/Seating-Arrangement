import pandas as pd
import os
import numpy as np
import csv
#f=open("f.csv","a",newline="")
#csvwrt=csv.writer(f)
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.worksheet.pagebreak import Break
wb = openpyxl.Workbook() 
sheet = wb.active 
df= pd.read_excel('Alternate Room Alotment.xlsx', sheet_name=[0])
num=int(input("Enter Number Of Seats In a Room: "))
n_row=int(input("Enter Number Of Rows In a Room: "))
n_col=num//n_row
room=list(df[0]['Room'].unique())
start_row=1
for r in room:
    l=[np.NAN]*(num)
    l2=[np.NAN]*(num//2)
    xx=list(pd.DataFrame(df[0][df[0]['Room']==r],columns=['ID','RN','Class'])['ID'])
    #num=len(xx)
    d=pd.DataFrame(df[0][df[0]['Room']==r],columns=['ID','RN','Class','Section'])
    '''row1=list(pd.DataFrame(df[0][df[0]['Room']==r],columns=['ID','RN','Class'])['ID'])[0:num//2]
    #rn1=list(pd.DataFrame(df[0][df[0]['Room']==r],columns=['ID','RN','Class'])['RN'])[0:num//2]
    row2=list(pd.DataFrame(df[0][df[0]['Room']==r],columns=['ID','RN','Class'])['ID'])[num//2:]
    #rn2=list(pd.DataFrame(df[0][df[0]['Room']==r],columns=['ID','RN','Class'])['RN'])[0:num//2:]
    #print(len(row1)+len(row2))'''
    for i in xx:
        if xx.index(i)<len(l):
            l[xx.index(i)]=str(i)+' Roll No. '+str(int(list(df[0][df[0]['ID']==i]['RN'])[0]))#"R.No"+str(r)
            #l1[row1.index(i)]=i+' R.No:'+str(rn1[row1.index(i)])
    '''for j in row2:
        if row2.index(j)<len(l2):
            #print(type(list(pd.DataFrame(df[2][df[2]['ID']==j],columns=['ID','RN'])['RN'])[0]))
            l2[row2.index(j)]=str(j)+'Roll No. '+str(int(list(df[0][df[0]['ID']==j]['RN'])[0]))
    #print(l1,l2)'''
    dict1={}
    for i in range(0,n_row):
        dict1['Row'+str(i+1)]=l[i*n_col:n_col*(i+1)]
    #print(dict1)
    #ddd={'Row1':l1[0:num//n_row],'Row3':l2[0:num//n_row],'Row2':l1[num//n_row:num//2],'Row4':l2[num//n_row:num//2]}
    #print(ddd)
    #dict1={'Row1':l1[0:num//4],'Row3':l2[0:num//4],'Row2':l1[num//4:num//2],'Row4':l2[num//4:num//2]}
    df1=pd.DataFrame(dict1,columns=dict1.keys(),index=np.arange(1,(num//n_row)+1))
    print(df1)
    sheet.merge_cells('A'+str(start_row)+':'+str(chr(64+n_row))+str(start_row))
    sheet.cell(row = start_row, column = 1).value='Room No-'+str(r)+" (PreBoard-Feb-Mar-22)"
    #sheet.cell(row=start_row,column=1).width=sheet.cell(row=start_row,column=1).value
    sheet.cell(row = start_row, column = 1).font=Font(size = 24,bold=True)
    sheet.cell(row = start_row, column = 1).alignment=Alignment(horizontal='center')
    start_row+=1
    for k in range(len(dict1)):
        sheet.cell(row=start_row,column=k+1).value='Row '+str(k+1)
        sheet.cell(row = start_row, column = k+1).alignment=Alignment(horizontal='center')
        sheet.cell(row = start_row, column = k+1).font=Font(size = 16,bold=True)
    '''sheet.cell(row=start_row,column=2).value='Row 2'
    sheet.cell(row = start_row, column = 2).alignment=Alignment(horizontal='center')
    sheet.cell(row = start_row, column = 2).font=Font(size = 16,bold=True)
    sheet.cell(row=start_row,column=3).value='Row 3'
    sheet.cell(row = start_row, column = 3).alignment=Alignment(horizontal='center')
    sheet.cell(row = start_row, column = 3).font=Font(size = 16,bold=True)
    sheet.cell(row=start_row,column=4).value='Row 4'
    sheet.cell(row = start_row, column = 4).alignment=Alignment(horizontal='center')
    sheet.cell(row = start_row, column = 4).font=Font(size = 16,bold=True)  '''  
    for i in range(1,len(df1)+1):
        start_row+=1
        for col in range(n_row):
            sheet.cell(row=start_row,column=col+1).value=df1.loc[i][col]
        '''sheet.cell(row=start_row,column=2).value=df1.loc[i][1]
        sheet.cell(row=start_row,column=3).value=df1.loc[i][2]
        sheet.cell(row=start_row,column=4).value=df1.loc[i][3]'''
    total=d.pivot_table(index=['Class','Section'],values=['ID'],aggfunc='count')
    start_row+=1
    sheet.cell(row=start_row,column=1).value='Class'
    sheet.cell(row = start_row, column = 1).font=Font(size = 13,bold=True)
    sheet.cell(row=start_row,column=2).value='Section'
    sheet.cell(row = start_row, column = 2).font=Font(size = 13,bold=True)
    sheet.cell(row=start_row,column=3).value='No Of Student'
    sheet.cell(row = start_row, column = 3).font=Font(size = 13,bold=True)
    for x in total.ID.iteritems():        
        start_row+=1
        sheet.cell(row=start_row,column=1).value=x[0][0]
        sheet.cell(row = start_row, column = 1).font=Font(size = 12,bold=True)
        sheet.cell(row=start_row,column=2).value=x[0][1]
        sheet.cell(row = start_row, column = 2).font=Font(size = 12,bold=True)
        sheet.cell(row=start_row,column=3).value=x[1]
        sheet.cell(row = start_row, column = 3).font=Font(size = 12,bold=True)
    
    #fred=sheet.pagebreak.Break(id=start_row)
    #sheet.pagebreak.PageBreak(brk=[fred])
    #sheet.append(fred)
    start_row=len(sheet['A'])+1
wb.save('Alternate Seating Plan.xlsx')       
print("Plan Generated  Successfully")
