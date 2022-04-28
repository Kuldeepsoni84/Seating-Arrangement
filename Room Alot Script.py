import pandas as pd
import os
import numpy as np
import csv
#f=open("f.csv","a",newline="")
#csvwrt=csv.writer(f)
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
x=int(input("enter No of Seat In a Room"))
wb = openpyxl.Workbook() 
sheet = wb.active 
df=pd.read_excel('X-XII Student List.xlsx',sheet_name=[0])
dfix=len(pd.DataFrame(df[0][df[0]["Group"]==1]))
dfx=len(pd.DataFrame(df[0][df[0]["Group"]==2]))
#dfxi=pd.DataFrame(df[0][df[0]["Class"]=="11th"])
#dfxii=pd.DataFrame(df[0][df[0]["Class"]=="12th"])
#global df_final
df_final=pd.DataFrame()
simple_df=pd.DataFrame()
alternate_df=pd.DataFrame()
dff=pd.DataFrame()
for i in range(0,max(dfix,dfx),x//2):
    first_df=pd.DataFrame(df[0][(df[0]["Group"]==1) ],
                         columns=['PR','Name','RN','Class','Section',"ID"])[i:i+x//2]
    #print(dfix_14)
    second_df=pd.DataFrame(df[0][(df[0]["Group"]==2 )],
                        columns=['PR','Name','RN','Class','Section',"ID"])[i:i+x//2]
    df_final=pd.concat([df_final,first_df,second_df])
    j,k=0,0
    for i in range(x):
        #arrange alternate
        if i%2==0:
            dff=pd.concat([dff,first_df[j:j+1]])
            j+=1
        else:
            dff=pd.concat([dff,second_df[k:k+1]])
            k+=1
                    
'''for i in range(0,max(len(dfix),len(dfx)),14):
    dfix_14=pd.DataFrame(df[0][(df[0]["Class"]=="9th") &(df[0]["Gender"]=="F")],columns=[
        'PR','Name','RN','Class','Section'])[i:i+14]
    second_df=pd.DataFrame(df[0][(df[0]["Class"]=="10th" )&(df[0]["Gender"]=="F")],columns=[
        'PR','Name','RN','Class','Section'])[i:i+14]
    df_final=pd.concat([df_final,dfix_14,second_df])'''
for i in range(0,len(df_final),x):
    df=pd.DataFrame(df_final[i:i+x])
    df["Room"]=(i//x)+1
    simple_df=pd.concat([simple_df,df])
for i in range(0,len(dff),x):
    df3=pd.DataFrame(dff[i:i+x])
    df3["Room"]=(i//x)+1
    alternate_df=pd.concat([alternate_df,df3])
#simple_df.to_excel("IX-XI Simple Room Alot2.xlsx",index = False)    
alternate_df.to_excel("Alternate Room Alotment.xlsx",index = False)
att_df=pd.read_excel('Alternate Room Alotment.xlsx', sheet_name=[0])
att_df[0].sort_values(by=['Class', 'Section','Name'])
room=list(att_df[0]['Room'].unique())
print(room)
wb = openpyxl.Workbook() 
sheet = wb.active
start_row=1
date=[]
for r in room:
    d=pd.DataFrame(att_df[0][att_df[0]['Room']==r],columns=['PR','Name','RN','Class','Section','ID','Room'])
    d['Name']=d['Name'].str.upper()
    d=d.sort_values(by=['Class', 'Section','Name'])
    d.reset_index(inplace = True)
    c_name='ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    col=columns=['PR','Name','RN','Class','Section','ID','25-02-22','26-02-22','28-02-22','01-03-22','02-03-22','03-03-22','04-03-22']
    sheet.merge_cells('A'+str(start_row)+':'+c_name[len(col)-1]+str(start_row))
    sheet.cell(row = start_row, column = 1).value='Room No-'+str(r)+" (PreBoard-Feb-Mar-22)"
    sheet.cell(row = start_row, column = 1).font=Font(size = 18,bold=True)
    sheet.cell(row = start_row, column = 1).alignment=Alignment(horizontal='center')
    start_row+=1
    
    k=1
    for c in col:
        sheet.cell(row=start_row,column=k).value=c
        sheet.cell(row = start_row, column = k).font=Font(size = 12,bold=True)
        sheet.cell(row = start_row, column = k).alignment=Alignment(horizontal='center')
        k+=1
    for i in range(0,len(d)):
        start_row+=1
        #print(len(sheet['A']),'start_row',start_row,"room_no",r)
        sheet.cell(row=start_row,column=1).value=d.loc[i][1]
        sheet.cell(row=start_row,column=2).value=d.loc[i][2]
        sheet.cell(row=start_row,column=3).value=d.loc[i][3]
        sheet.cell(row=start_row,column=4).value=d.loc[i][4]
        sheet.cell(row=start_row,column=5).value=d.loc[i][5]
        sheet.cell(row=start_row,column=6).value=d.loc[i][6]
        #sheet.cell(row=start_row,column=7).value=d.loc[i][7]
    start_row+=1
    sheet.merge_cells('A'+str(start_row)+':F'+str(start_row))
    sheet.cell(row=start_row,column=1).value='Total Present'
    sheet.cell(row = start_row, column = 1).font=Font(size = 10,bold=True)
    sheet.cell(row = start_row, column = 1).alignment=Alignment(horizontal='right')
    start_row+=1
    sheet.merge_cells('A'+str(start_row)+':F'+str(start_row))
    sheet.cell(row=start_row,column=1).value='Total Absent'
    sheet.cell(row = start_row, column = 1).font=Font(size = 10,bold=True)
    sheet.cell(row = start_row, column = 1).alignment=Alignment(horizontal='right')
    start_row+=1
    sheet.merge_cells('A'+str(start_row)+':F'+str(start_row))
    sheet.cell(row=start_row,column=1).value='Signature'
    sheet.cell(row = start_row, column = 1).font=Font(size = 10,bold=True)
    sheet.cell(row = start_row, column = 1).alignment=Alignment(horizontal='right')
    start_row=len(sheet['A'])+1
    
wb.save('Attendance sheet.xlsx') 
        
    
    

    

